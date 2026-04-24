"""
Clinical Operations System (COS) — pre-launch financial workbook [DRAFT_OP_V5].
Tier matrix, director hybrid (65%×Senior rate), adjustment reserve, benchmark ranges, print dossier.
"""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils.protection import hash_password


H = " [DRAFT_OP_V5]"
# Sheet protection (review / publisher handoff). Replace before external distribution if required.
WORKBOOK_SHEET_PASSWORD = "COS-2026-QC-REVIEW"
FILE_INFRASTRUCTURE_ASSETS = "INFRASTRUCTURE_ASSETS.csv"
FILE_INITIAL_CAPITAL_VALUATION = "INITIAL_CAPITAL_VALUATION.csv"
BLUE_INPUT = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
GRAY_GHOST = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
GREEN_CF = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_CF = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FONT = Font(bold=True)
GHOST_FONT = Font(italic=True, color="555555", size=10)
VALUE_FONT = Font(bold=True)
PENDING_NOTE = "[PENDING_DATA]"
ST_VERIFIED = "[VERIFIED]"
ST_PENDING = "[PENDING_DATA]"
DASHBOARD_ADJ_CELL = "N38"


def _draft_comment(text: str) -> Comment:
    return Comment(f"DRAFT_OP_V5: {text}", "The Architect")


def _load_infrastructure_line_items(base: Path) -> list[tuple[str, float, float, float]]:
    """Return (description, hours, rate, direct_total) per CSV row; direct_total used when hours=0."""
    p = base / FILE_INFRASTRUCTURE_ASSETS
    if not p.is_file():
        p = base / "SHANE_ASSETS.csv"
    if not p.is_file():
        return []
    rows: list[tuple[str, float, float, float]] = []
    with p.open(newline="", encoding="utf-8") as f:
        r = csv.DictReader(f)
        if not r.fieldnames:
            return []
        for row in r:
            desc = (row.get("Description") or "").strip() or "Line item"
            try:
                h = float(row.get("Hours") or 0)
            except (TypeError, ValueError):
                h = 0.0
            try:
                rate = float(row.get("Market_Rate_CAD") or 125)
            except (TypeError, ValueError):
                rate = 125.0
            raw_tv = row.get("Total_Value_CAD") or row.get("Total_Value") or 0
            try:
                tv = float(raw_tv)
            except (TypeError, ValueError):
                tv = 0.0
            rows.append((desc, h, rate, tv))
    return rows


def _load_infrastructure_assets_total(base: Path) -> float:
    p = base / FILE_INFRASTRUCTURE_ASSETS
    if not p.is_file():
        p = base / "SHANE_ASSETS.csv"
    if not p.is_file():
        return 0.0
    t = 0.0
    with p.open(newline="", encoding="utf-8") as f:
        r = csv.DictReader(f)
        if not r.fieldnames or (
            "Total_Value" not in r.fieldnames and "Total_Value_CAD" not in r.fieldnames
        ):
            return 0.0
        for row in r:
            raw = row.get("Total_Value") or row.get("Total_Value_CAD") or 0
            try:
                t += float(raw)
            except (TypeError, ValueError):
                pass
    return t


def _load_initial_capital_valuation_total(base: Path) -> float:
    """INITIAL_CAPITAL_VALUATION.csv — audited hours + subscription rollup."""
    p = base / FILE_INITIAL_CAPITAL_VALUATION
    if not p.is_file():
        p = base / "SHANE_VALUATION.csv"
    if not p.is_file():
        return 0.0
    t = 0.0
    with p.open(newline="", encoding="utf-8") as f:
        r = csv.DictReader(f)
        if not r.fieldnames or "Total_Value" not in r.fieldnames:
            return 0.0
        for row in r:
            raw = row.get("Total_Value") or row.get("Total_Value_CAD") or 0
            try:
                t += float(raw)
            except (TypeError, ValueError):
                pass
    return t


def _export_model_sheet_csv(wb: Workbook, path: Path) -> None:
    """Publisher: flat MODEL snapshot for audit (MODEL.csv)."""
    ws = wb["MODEL"]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            w.writerow(["" if c is None else c for c in row])


def _lock_all_cells(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=True)


def _unlock_cells(ws, addresses: list[str]) -> None:
    for addr in addresses:
        ws[addr].protection = Protection(locked=False)


def build_workbook(output_path: str | Path | None = None) -> Path:
    base = Path(__file__).resolve().parent.parent
    out = Path(output_path) if output_path else base / "output" / "ANALISIS_FINANCIERO_PRE_LAUNCH.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.properties.title = "Clinical Operations System (COS) — Pre-Launch Financial Model"
    wb.properties.creator = "The Architect"

    fmt_int = "0"
    fmt_money = "#,##0"
    fmt_pct = "0.00"
    fmt_tax = "0.00"

    # --- INPUTS row constants (1-based) — must match inputs_spec order below ---
    INPUTS_CLINIC_PCT_ROW = 2
    INPUTS_UTIL_ROW = 3
    INPUTS_WEEKS_YR_ROW = 4
    INPUTS_WEEKS_FACTOR_ROW = 5
    INPUTS_PARTNER_DRAW_ROW = 6
    INPUTS_DIR_ADMIN_HRS_ROW = 7
    INPUTS_DIR_ADMIN_RATE_ROW = 8
    INPUTS_DIR_CLIN_SESS_ROW = 9
    INPUTS_DIR_BILL_RATE_ROW = 10
    INPUTS_DIR_CLIN_PAY_ROW = 11
    INPUTS_PS_MIN_ROW = 12
    INPUTS_SUP_HOURS_ROW = 13
    INPUTS_SUP_RATE_ROW = 14
    INPUTS_ADMIN_FLAT_ROW = 15
    INPUTS_DIR_PCT_ROW = 16
    INPUTS_DIR_SESS_FEE_ROW = 17
    INPUTS_TAX_ROW = 19
    INPUTS_PLANNING_NP_ROW = 20
    INPUTS_INVESTOR_NET_ROW = 21
    INPUTS_PARTNER_L_NEED_ROW = 22
    INPUTS_DIVIDEND_ROW = 23
    INPUTS_IMPL_DEBT_ROW = 25
    INPUTS_IMPL_CAP_ROW = 26
    INPUTS_LEGAL_GHOST_ROW = 28
    INPUTS_LEGAL_ACTUAL_ROW = 29
    INPUTS_MKT_GHOST_ROW = 30
    INPUTS_MKT_MO_ROW = 31
    INPUTS_JANE_GHOST_ROW = 32
    INPUTS_JANE_MO_ROW = 33
    INPUTS_INS_GHOST_ROW = 38
    INPUTS_INS_MO_ROW = 39
    INPUTS_ADJUSTMENT_RESERVE_ROW = 41
    INPUTS_INITIAL_TOTAL_ROW = 42

    inputs_spec: list[tuple[str | None, object | None, str | None]] = [
        (f"VARIABLE{H}", f"VALUE{H}", f"STATUS{H}"),
        (f"Clinic %{H}", 0.33, ST_VERIFIED),
        (f"Utilization (Path B ref.){H}", 0.80, ST_PENDING),
        (f"Weeks/year{H}", 48, ST_PENDING),
        (f"Weeks/month factor (sessions equiv.){H}", "=52/12", None),
        (f"Partner_L_Draw — fixed target (CAD/mo){H}", 2000, ST_VERIFIED),
        (f"Director — admin hours/month{H}", 20, ST_PENDING),
        (f"Director_Admin_Rate ($/hr){H}", 75, ST_VERIFIED),
        (f"Director — clinical sessions/month{H}", 0, ST_PENDING),
        (f"Director billing rate — Senior (CAD/session){H}", 190, ST_VERIFIED),
        (f"Director clinical pay — 65% × Senior rate (=0.65×billing){H}", "=0.65*B10", None),
        (f"Min roster count (activate fees){H}", 1, ST_PENDING),
        (f"Supervision — hours/month{H}", 0, ST_PENDING),
        (f"Supervision — rate (CAD/hr){H}", 150, ST_VERIFIED),
        (f"Admin / platform flat fee (CAD/mo){H}", 0, ST_PENDING),
        (f"Director legacy — % of roster gross{H}", 0, ST_PENDING),
        (f"Director legacy — $/session (roster volume){H}", 0, ST_PENDING),
        (None, None, None),
        (f"Tax rate — clinic{H}", 0.30, ST_PENDING),
        (f"Planning — NP benchmark (not an expense){H}", 0, ST_PENDING),
        (f"Investor net target (CAD/mo){H}", 0, ST_PENDING),
        (f"Partner_L monthly need (reference){H}", 0, ST_PENDING),
        (f"Dividend payout ratio (of net profit){H}", 0.25, ST_PENDING),
        (None, None, None),
        (
            f"Initial capital — line-item audit (→ INFRASTRUCTURE_ASSETS){H}",
            0,
            ST_VERIFIED,
        ),
        (
            f"Initial capital — valuation rollup (→ CAPEX_SUMMARY){H}",
            0,
            ST_VERIFIED,
        ),
        (None, None, None),
        (
            f"Legal — incorporation — RANGE (ghost){H}",
            "Market Benchmark [QC_2026_DATA]: [$600 – $5,000] QC incorporation (reference band)",
            None,
        ),
        (f"Legal — incorporation — ACTUAL (CAD){H}", 600, ST_VERIFIED),
        (
            f"Marketing (monthly) — RANGE (ghost){H}",
            "Market Benchmark [QC_2026_DATA]: [$500 – $2,500] Google Ads (reference band)",
            None,
        ),
        (f"Marketing (monthly) — ACTUAL (CAD){H}", 500, ST_VERIFIED),
        (
            f"Software (Jane App) — RANGE (ghost){H}",
            "Market Benchmark [QC_2026_DATA]: [$79 – $109] Jane App (reference band)",
            None,
        ),
        (f"Software (Jane App) — ACTUAL (CAD/mo){H}", 109, ST_VERIFIED),
        (f"Initial investment — Marketing launch (one-time){H}", 0, ST_PENDING),
        (f"Initial investment — Software licenses{H}", 0, ST_PENDING),
        (f"Municipal license / business registration (annual CAD){H}", 300, ST_PENDING),
        (f"Professional liability — Social Work (annual CAD){H}", 400, ST_PENDING),
        (
            f"Professional insurance (QC psych) — RANGE (ghost){H}",
            "Market Benchmark [QC_2026_DATA]: [$50 – $100]/mo (reference band)",
            None,
        ),
        (f"Professional insurance (QC psych) — ACTUAL (CAD/mo){H}", 65, ST_VERIFIED),
        (f"Monthly — Domain / Hosting / Phone{H}", 85, ST_VERIFIED),
        (
            f"ADJUSTMENT_RESERVE (→ DASHBOARD Input_Variables){H}",
            0,
            ST_VERIFIED,
        ),
        (f"Initial investment total{H}", None, None),
    ]

    ws_in = wb.active
    ws_in.title = "INPUTS"

    for r, (a, b, st) in enumerate(inputs_spec, start=1):
        if a is not None:
            ws_in[f"A{r}"] = a
        if b is not None:
            ws_in[f"B{r}"] = b
        if st is not None:
            ws_in[f"C{r}"] = st
        if b is None and a is None:
            continue
        if isinstance(b, str) and b.startswith("="):
            ws_in[f"B{r}"].fill = PatternFill()
        elif b is not None and st is not None:
            ws_in[f"B{r}"].fill = BLUE_INPUT

    for r in (
        INPUTS_LEGAL_GHOST_ROW,
        INPUTS_MKT_GHOST_ROW,
        INPUTS_JANE_GHOST_ROW,
        INPUTS_INS_GHOST_ROW,
    ):
        ws_in[f"A{r}"].fill = GRAY_GHOST
        ws_in[f"B{r}"].fill = GRAY_GHOST
        ws_in[f"A{r}"].font = GHOST_FONT
        ws_in[f"B{r}"].font = GHOST_FONT
    for r in (INPUTS_LEGAL_ACTUAL_ROW, INPUTS_MKT_MO_ROW, INPUTS_JANE_MO_ROW, INPUTS_INS_MO_ROW):
        ws_in[f"A{r}"].font = VALUE_FONT

    _ghost_pair_ht = 24.0
    for r in (
        INPUTS_MKT_GHOST_ROW,
        INPUTS_MKT_MO_ROW,
        INPUTS_JANE_GHOST_ROW,
        INPUTS_JANE_MO_ROW,
        INPUTS_INS_GHOST_ROW,
        INPUTS_INS_MO_ROW,
    ):
        ws_in.row_dimensions[r].height = _ghost_pair_ht

    ws_in[f"B{INPUTS_INITIAL_TOTAL_ROW}"] = (
        f"=SUM(B{INPUTS_LEGAL_ACTUAL_ROW},B{INPUTS_LEGAL_ACTUAL_ROW + 5},B{INPUTS_LEGAL_ACTUAL_ROW + 6})"
    )
    ws_in[f"B{INPUTS_INITIAL_TOTAL_ROW}"].number_format = fmt_money
    ws_in[f"B{INPUTS_INITIAL_TOTAL_ROW}"].comment = _draft_comment("Legal actual + marketing launch one-time + software init.")

    for r, (_, b, st) in enumerate(inputs_spec, start=1):
        if b is None or st is None or (isinstance(b, str) and str(b).startswith("=")):
            if isinstance(b, str) and str(b).startswith("="):
                ws_in[f"B{r}"].number_format = "0.0000"
            continue
        if r in (INPUTS_CLINIC_PCT_ROW, INPUTS_UTIL_ROW):
            ws_in[f"B{r}"].number_format = fmt_pct
        elif r == INPUTS_WEEKS_YR_ROW:
            ws_in[f"B{r}"].number_format = fmt_int
        elif r in (
            INPUTS_PARTNER_DRAW_ROW,
            INPUTS_DIR_ADMIN_HRS_ROW,
            INPUTS_DIR_CLIN_SESS_ROW,
            INPUTS_PS_MIN_ROW,
            INPUTS_SUP_HOURS_ROW,
        ):
            ws_in[f"B{r}"].number_format = fmt_int if r != INPUTS_PARTNER_DRAW_ROW else fmt_money
        elif r in (
            INPUTS_DIR_ADMIN_RATE_ROW,
            INPUTS_DIR_BILL_RATE_ROW,
            INPUTS_SUP_RATE_ROW,
            INPUTS_ADMIN_FLAT_ROW,
        ):
            ws_in[f"B{r}"].number_format = fmt_money
        elif r in (INPUTS_DIR_PCT_ROW,):
            ws_in[f"B{r}"].number_format = fmt_pct
        elif r in (INPUTS_DIR_SESS_FEE_ROW,):
            ws_in[f"B{r}"].number_format = fmt_money
        elif r == INPUTS_TAX_ROW:
            ws_in[f"B{r}"].number_format = fmt_tax
        elif r in (INPUTS_PLANNING_NP_ROW, INPUTS_INVESTOR_NET_ROW, INPUTS_PARTNER_L_NEED_ROW):
            ws_in[f"B{r}"].number_format = fmt_money
        elif r == INPUTS_DIVIDEND_ROW:
            ws_in[f"B{r}"].number_format = fmt_pct
        elif r in (INPUTS_IMPL_DEBT_ROW, INPUTS_IMPL_CAP_ROW):
            ws_in[f"B{r}"].number_format = fmt_money
        elif r in (INPUTS_LEGAL_ACTUAL_ROW, INPUTS_MKT_MO_ROW, INPUTS_JANE_MO_ROW):
            ws_in[f"B{r}"].number_format = fmt_money
        elif r in (
            INPUTS_LEGAL_ACTUAL_ROW + 5,
            INPUTS_LEGAL_ACTUAL_ROW + 6,
            36,
            37,
            39,
            40,
            INPUTS_ADJUSTMENT_RESERVE_ROW,
        ):
            ws_in[f"B{r}"].number_format = fmt_money

    ws_in["A1"].font = HEADER_FONT
    ws_in["B1"].font = HEADER_FONT
    ws_in["C1"].font = HEADER_FONT
    ws_in.column_dimensions["A"].width = 54
    ws_in.column_dimensions["B"].width = 22
    ws_in.column_dimensions["C"].width = 16

    municipal_inputs_row = 36
    sw_inputs_row = 37
    ins_mo_row = INPUTS_INS_MO_ROW
    dhp_mo_row = 40

    weeks_f = f"INPUTS!$B${INPUTS_WEEKS_FACTOR_ROW}"
    clinic_ref = f"INPUTS!$B${INPUTS_CLINIC_PCT_ROW}"

    # --- TIERS (Intern / Junior / Senior benchmark rates) ---
    ws_tier = wb.create_sheet("TIERS", 1)
    ws_tier["A1"] = f"Tier{H}"
    ws_tier["B1"] = f"Rate_Per_Session (CAD){H}"
    for c in ("A1", "B1"):
        ws_tier[c].font = HEADER_FONT
    tier_rows = [("Intern", 120), ("Junior", 150), ("Senior", 190)]
    for i, (tname, rate) in enumerate(tier_rows, start=2):
        ws_tier[f"A{i}"] = f"{tname}{H}"
        ws_tier[f"B{i}"] = rate
        ws_tier[f"A{i}"].fill = BLUE_INPUT
        ws_tier[f"B{i}"].fill = BLUE_INPUT
        ws_tier[f"B{i}"].number_format = fmt_money
    ws_tier.column_dimensions["A"].width = 14
    ws_tier.column_dimensions["B"].width = 22

    # --- ROSTER ---
    ws_ro = wb.create_sheet("ROSTER", 2)
    ws_ro["A1"] = f"Therapist_ID{H}"
    ws_ro["B1"] = f"Tier (Intern / Jr / Sr){H}"
    ws_ro["C1"] = f"Weekly_Hours{H}"
    ws_ro["D1"] = f"Rate_Per_Session (CAD){H}"
    ws_ro["E1"] = f"Monthly clinic share{H}"
    ws_ro["F1"] = f"Monthly gross billing (ref.){H}"
    for c in ("A1", "B1", "C1", "D1", "E1", "F1"):
        ws_ro[c].font = HEADER_FONT
    roster_first, roster_last = 2, 21
    for r in range(roster_first, roster_last + 1):
        ws_ro[f"E{r}"] = f"=$C{r}*4*$D{r}*{clinic_ref}"
        ws_ro[f"F{r}"] = f"=$C{r}*4*$D{r}"
        for col in ("A", "B", "C", "D"):
            ws_ro[f"{col}{r}"].fill = BLUE_INPUT
        ws_ro[f"C{r}"].number_format = fmt_int
        ws_ro[f"D{r}"].number_format = fmt_money
        ws_ro[f"E{r}"].number_format = fmt_money
        ws_ro[f"F{r}"].number_format = fmt_money

    roster_sum_row = roster_last + 1
    ws_ro[f"A{roster_sum_row}"] = f"TOTAL{H}"
    ws_ro[f"A{roster_sum_row}"].font = HEADER_FONT
    ws_ro[f"E{roster_sum_row}"] = f"=SUM(E{roster_first}:E{roster_last})"
    ws_ro[f"F{roster_sum_row}"] = f"=SUM(F{roster_first}:F{roster_last})"
    ws_ro[f"E{roster_sum_row}"].number_format = fmt_money
    ws_ro[f"F{roster_sum_row}"].number_format = fmt_money
    for col, w in (("A", 14), ("B", 18), ("C", 14), ("D", 16), ("E", 18), ("F", 22)):
        ws_ro.column_dimensions[col].width = w

    ROSTER_E_SUM = f"ROSTER!$E${roster_sum_row}"
    ROSTER_F_SUM = f"ROSTER!$F${roster_sum_row}"
    ROSTER_COUNT = "COUNTA(ROSTER!$A$2:$A$21)"

    # --- OPERATING_COSTS ---
    ws_oc = wb.create_sheet("OPERATING_COSTS", 3)
    ws_oc["A1"] = f"Operating costs — detail{H}"
    ws_oc["A1"].font = Font(bold=True, size=12)
    ws_oc["A3"] = f"Category{H}"
    ws_oc["B3"] = f"Monthly CAD{H}"
    ws_oc["C3"] = f"Notes{H} {PENDING_NOTE}"
    ws_oc["D3"] = f"In Path B (light){H}"
    for c in ("A3", "B3", "C3", "D3"):
        ws_oc[c].font = HEADER_FONT

    oc_lines = [
        (f"Professional insurance — QC (psych){H}", "N"),
        (f"Regulatory compliance — OPQ{H}", "N"),
        (f"Software — base stack (excl. Jane){H}", "Y"),
        (f"Jane App (Compliance){H}", "Y"),
        (f"Automated marketing (base){H}", "Y"),
        (f"Domain / Hosting / Phone{H}", "Y"),
        (f"Other operating{H}", "Y"),
        (f"Municipal license / business registration{H}", "Y"),
        (f"Professional liability — Social Work{H}", "Y"),
    ]
    oc_first = 4
    oc_last = oc_first + len(oc_lines) - 1

    for i, (label, path_b) in enumerate(oc_lines):
        r = oc_first + i
        ws_oc[f"A{r}"] = label
        ws_oc[f"D{r}"] = path_b
        ws_oc[f"C{r}"] = "—"
        if "Professional insurance" in label and "Municipal" not in label:
            ws_oc[f"B{r}"] = f"=INPUTS!$B${ins_mo_row}"
        elif "Jane App" in label:
            ws_oc[f"B{r}"] = f"=INPUTS!$B${INPUTS_JANE_MO_ROW}"
        elif "marketing" in label.lower():
            ws_oc[f"B{r}"] = f"=INPUTS!$B${INPUTS_MKT_MO_ROW}"
        elif "Domain" in label or "Hosting" in label:
            ws_oc[f"B{r}"] = f"=INPUTS!$B${dhp_mo_row}"
        elif "Municipal" in label:
            ws_oc[f"B{r}"] = f"=INPUTS!$B${municipal_inputs_row}/12"
        elif "Social Work" in label:
            ws_oc[f"B{r}"] = f"=INPUTS!$B${sw_inputs_row}/12"
        else:
            ws_oc[f"B{r}"] = 0
        ws_oc[f"B{r}"].fill = BLUE_INPUT
        ws_oc[f"B{r}"].number_format = fmt_money

    oc_total_row = oc_last + 2
    ws_oc[f"A{oc_total_row}"] = f"Total monthly (all lines){H}"
    ws_oc[f"B{oc_total_row}"] = f"=SUM(B{oc_first}:B{oc_last})"
    ws_oc[f"B{oc_total_row}"].number_format = fmt_money
    ws_oc[f"A{oc_total_row}"].font = HEADER_FONT

    oc_path_b_row = oc_total_row + 1
    ws_oc[f"A{oc_path_b_row}"] = f"Path B — monthly opex (excl. QC ins. + OPQ){H}"
    ws_oc[f"B{oc_path_b_row}"] = f'=SUMIF(D{oc_first}:D{oc_last},"Y",B{oc_first}:B{oc_last})'
    ws_oc[f"B{oc_path_b_row}"].number_format = fmt_money
    ws_oc[f"A{oc_path_b_row}"].font = Font(bold=True, italic=True)

    for col, w in (("A", 50), ("B", 14), ("C", 28), ("D", 18)):
        ws_oc.column_dimensions[col].width = w

    opex_total = f"OPERATING_COSTS!$B${oc_total_row}"
    opex_path_b = f"OPERATING_COSTS!$B${oc_path_b_row}"

    # --- INFRASTRUCTURE_ASSETS (CSV line items → CAPEX baseline) ---
    ws_av = wb.create_sheet("INFRASTRUCTURE_ASSETS", 4)
    ws_av["A1"] = f"Infrastructure assets — line items ({FILE_INFRASTRUCTURE_ASSETS}){H}"
    ws_av["A1"].font = Font(bold=True, size=12)
    ws_av["A3"] = f"Line item{H}"
    ws_av["B3"] = f"Hours{H}"
    ws_av["C3"] = f"Reference rate ($/hr){H}"
    ws_av["D3"] = f"Value (CAD){H}"
    for c in ("A3", "B3", "C3", "D3"):
        ws_av[c].font = HEADER_FONT
    av_items = _load_infrastructure_line_items(base)
    if not av_items:
        av_items = [
            ("[Placeholder — import CSV]", 0.0, 125.0, 0.0),
        ]
    av_first = 4
    av_last_data = av_first + len(av_items) - 1
    for i, (desc, h, rate, tv) in enumerate(av_items):
        r = av_first + i
        ws_av[f"A{r}"] = f"{desc}{H}"
        ws_av[f"B{r}"] = h
        ws_av[f"C{r}"] = rate
        if abs(h) > 1e-12:
            ws_av[f"D{r}"] = f"=B{r}*C{r}"
        else:
            ws_av[f"D{r}"] = tv
        for col in ("A", "B", "C", "D"):
            ws_av[f"{col}{r}"].fill = BLUE_INPUT
        ws_av[f"B{r}"].number_format = fmt_int if abs(h - round(h)) < 1e-9 else "0.00"
        ws_av[f"C{r}"].number_format = fmt_money
        ws_av[f"D{r}"].number_format = fmt_money
    av_total_row = av_last_data + 1
    ws_av[f"A{av_total_row}"] = f"Total — line-item value (CAD){H}"
    ws_av[f"B{av_total_row}"] = f"=SUM(D{av_first}:D{av_last_data})"
    ws_av[f"B{av_total_row}"].number_format = fmt_money
    ws_av[f"A{av_total_row}"].font = HEADER_FONT
    for col, w in (("A", 44), ("B", 12), ("C", 16), ("D", 18)):
        ws_av.column_dimensions[col].width = w
    ASSET_SWEAT_TOTAL = f"INFRASTRUCTURE_ASSETS!$B${av_total_row}"

    # --- CAPEX_SUMMARY ---
    ws_capex = wb.create_sheet("CAPEX_SUMMARY", 5)
    ws_capex["A1"] = f"CAPEX_SUMMARY{H}"
    ws_capex["A1"].font = Font(bold=True, size=12)
    ws_capex["A3"] = "Total Initial Valuation (CAD)"
    ws_capex["B3"] = f"=INFRASTRUCTURE_ASSETS!$B${av_total_row}"
    ws_capex["A4"] = "Architect's System Contribution"
    ws_capex["B4"] = "=B3"
    for addr in ("B3", "B4"):
        ws_capex[addr].number_format = fmt_money
        ws_capex[addr].font = Font(bold=True)
    ws_capex.column_dimensions["A"].width = 40
    ws_capex.column_dimensions["B"].width = 22

    ws_in[f"B{INPUTS_IMPL_DEBT_ROW}"] = f"=INFRASTRUCTURE_ASSETS!$B${av_total_row}"
    ws_in[f"B{INPUTS_IMPL_CAP_ROW}"] = "=CAPEX_SUMMARY!$B$3"
    ws_in[f"B{INPUTS_IMPL_DEBT_ROW}"].number_format = fmt_money
    ws_in[f"B{INPUTS_IMPL_CAP_ROW}"].number_format = fmt_money
    ws_in.merge_cells("A43:C43")
    ws_in["A43"] = (
        "Footnote: [PENDING_DATA] = operational inputs pending user entry. "
        "Initial capital (B25–B26) references INFRASTRUCTURE_ASSETS and CAPEX_SUMMARY."
    )
    ws_in["A43"].font = Font(size=9, italic=True, color="6B7280")
    ws_in["A43"].alignment = Alignment(wrap_text=True, vertical="top")

    # --- PARTNERSHIP_STRUCTURE ---
    ws_cap = wb.create_sheet("PARTNERSHIP_STRUCTURE", 6)
    ws_cap["A1"] = f"Partnership structure — ownership{H}"
    ws_cap["A1"].font = Font(bold=True, size=12)
    ws_cap["A2"] = f"Post-money valuation (CAD){H}"
    ws_cap["B2"] = 0
    ws_cap["B2"].fill = BLUE_INPUT
    ws_cap["B2"].number_format = fmt_money
    hdr = 4
    for col, lab in enumerate(
        [f"Party{H}", f"Capital ($){H}", f"Sweat hrs{H}", f"Rate ($){H}", f"Contribution{H}", f"% Own{H}"],
        start=1,
    ):
        ws_cap.cell(row=hdr, column=col, value=lab).font = HEADER_FONT
    parties = [
        (f"Principal A (infrastructure){H}", 0, 0, 0),
        (f"Partner L{H}", 0, 0, 0),
        (f"Principal C{H}", 0, 0, 0),
        (f"Director (res.){H}", 0, 0, 0),
    ]
    cap_first = 5
    cap_last = cap_first + len(parties) - 1
    for i, (name, cap, hrs, rate) in enumerate(parties):
        r = cap_first + i
        ws_cap[f"A{r}"] = name
        ws_cap[f"B{r}"] = cap
        ws_cap[f"C{r}"] = hrs
        ws_cap[f"D{r}"] = rate
        if i == 0:
            ws_cap[f"E{r}"] = f"=B{r}+C{r}*D{r}+{ASSET_SWEAT_TOTAL}"
        else:
            ws_cap[f"E{r}"] = f"=B{r}+C{r}*D{r}"
        ws_cap[f"F{r}"] = (
            f"=IF(SUM($E${cap_first}:$E${cap_last})=0,0,E{r}/SUM($E${cap_first}:$E${cap_last}))"
        )
        for col in ("A", "B", "C", "D"):
            ws_cap[f"{col}{r}"].fill = BLUE_INPUT
        ws_cap[f"B{r}"].number_format = fmt_money
        ws_cap[f"D{r}"].number_format = fmt_money
        ws_cap[f"E{r}"].number_format = fmt_money
        ws_cap[f"F{r}"].number_format = "0.00%"
    ws_cap["A13"] = f"Total{H}"
    ws_cap["B13"] = f"=SUM(E{cap_first}:E{cap_last})"
    ws_cap["A14"] = f"Δ vs post-money{H}"
    ws_cap["B14"] = "=B13-B2"
    for addr in ("B13", "B14"):
        ws_cap[addr].number_format = fmt_money
    CAP_PRINCIPAL_A_ROW = cap_first

    director_clinic_rev = (
        f"INPUTS!$B${INPUTS_DIR_CLIN_SESS_ROW}*INPUTS!$B${INPUTS_DIR_BILL_RATE_ROW}*{clinic_ref}"
    )
    director_hybrid_pay = (
        f"INPUTS!$B${INPUTS_DIR_ADMIN_HRS_ROW}*INPUTS!$B${INPUTS_DIR_ADMIN_RATE_ROW}+"
        f"INPUTS!$B${INPUTS_DIR_CLIN_SESS_ROW}*INPUTS!$B${INPUTS_DIR_CLIN_PAY_ROW}"
    )
    legacy_fee = (
        f"IF({ROSTER_COUNT}<INPUTS!$B${INPUTS_PS_MIN_ROW},0,"
        f"IF(INPUTS!$B${INPUTS_DIR_PCT_ROW}>0,{ROSTER_F_SUM}*INPUTS!$B${INPUTS_DIR_PCT_ROW},"
        f"IF(INPUTS!$B${INPUTS_DIR_SESS_FEE_ROW}>0,SUM(ROSTER!$C$2:$C$21)*4*INPUTS!$B${INPUTS_DIR_SESS_FEE_ROW},0)))"
    )
    admin_fee = (
        f"IF({ROSTER_COUNT}>=INPUTS!$B${INPUTS_PS_MIN_ROW},INPUTS!$B${INPUTS_ADMIN_FLAT_ROW},0)"
    )
    sup_cost = f"INPUTS!$B${INPUTS_SUP_HOURS_ROW}*INPUTS!$B${INPUTS_SUP_RATE_ROW}"

    operating_total = f"({opex_total}+{admin_fee}+{sup_cost}+{legacy_fee}+{director_hybrid_pay})"

    # --- MODEL (Director revenue first — solvency anchor; adjustment reserve last) ---
    ws_m = wb.create_sheet("MODEL", 7)
    ws_m["A1"] = f"METRIC{H}"
    ws_m["B1"] = f"VALUE{H}"
    ws_m["A1"].font = HEADER_FONT
    ws_m["B1"].font = HEADER_FONT

    ws_m["A3"] = f"Director — clinic share revenue (solvency anchor){H}"
    ws_m["B3"] = f"={director_clinic_rev}"
    ws_m["A4"] = f"Roster — clinic share{H}"
    ws_m["B4"] = f"={ROSTER_E_SUM}"
    ws_m["A5"] = f"Total clinic margin{H}"
    ws_m["B5"] = "=B3+B4"
    ws_m["A6"] = f"Operating_Costs (incl. director hybrid pay){H}"
    ws_m["B6"] = f"={operating_total}"
    ws_m["A7"] = f"Partner_L_Draw (fixed){H}"
    ws_m["B7"] = f"=INPUTS!$B${INPUTS_PARTNER_DRAW_ROW}"
    ws_m["A8"] = f"Pre-tax (before adjustment reserve){H}"
    ws_m["B8"] = "=B5-B6-B7"
    ws_m["A9"] = f"Tax{H}"
    ws_m["B9"] = f"=IF(B8>0,B8*INPUTS!$B${INPUTS_TAX_ROW},0)"
    ws_m["A10"] = f"Net after tax (before ADJUSTMENT_RESERVE){H}"
    ws_m["B10"] = "=B8-B9"
    ws_m["A11"] = f"ADJUSTMENT_RESERVE (INPUTS — last lever){H}"
    ws_m["B11"] = f"=INPUTS!$B${INPUTS_ADJUSTMENT_RESERVE_ROW}"
    ws_m["A12"] = f"Net profit (final){H}"
    ws_m["B12"] = "=B10+B11"
    ws_m["A13"] = f"Alert — equity redemption{H}"
    ws_m["B13"] = (
        f"=IF(B12<INPUTS!$B${INPUTS_PARTNER_DRAW_ROW},"
        f"\"EQUITY REDEMPTION REQUIRED\",\"\")"
    )
    ws_m["A14"] = f"Validation — solvency (ops + draw > margin){H}"
    ws_m["B14"] = f"=IF(B6+B7>B5,\"WARNING: CLINIC INSOLVENT\",\"\")"

    ws_m["A16"] = f"Reference — therapist-side (1 − Clinic %) × gross{H}"
    ws_m["B16"] = f"={ROSTER_F_SUM}*(1-{clinic_ref})"

    ws_m["A17"] = f"Break-even — roster gross (approx.){H}"
    ws_m["B17"] = (
        f"=IF({ROSTER_F_SUM}<=0,0,IF({clinic_ref}<=0,0,"
        f"({opex_total}+{admin_fee}+{sup_cost}+{legacy_fee})/{clinic_ref}))"
    )

    ws_m["A18"] = f"Dividend pool{H}"
    ws_m["B18"] = f"=IF(B12>0,B12*INPUTS!$B${INPUTS_DIVIDEND_ROW},0)"

    inv_start = 19
    inv_labels = ["Principal A", "Partner L", "Principal C", "Director (reserved)"]
    principal_a_div_row = inv_start
    for i, lab in enumerate(inv_labels):
        r = inv_start + i
        ws_m[f"A{r}"] = f"Dividend — {lab}{H}"
        ws_m[f"B{r}"] = f"=$B$18*PARTNERSHIP_STRUCTURE!$F${CAP_PRINCIPAL_A_ROW + i}"

    chk_r = inv_start + 4
    ws_m[f"A{chk_r}"] = f"Check (pool){H}"
    ws_m[f"B{chk_r}"] = f"=B18-SUM(B{inv_start}:B{inv_start + 3})"

    pay_row = chk_r + 1
    ws_m[f"A{pay_row}"] = f"Payback (months){H}"
    ws_m[f"B{pay_row}"] = (
        f'=IF(AND(B12>0,INPUTS!$B${INPUTS_INITIAL_TOTAL_ROW}>0),'
        f'INPUTS!$B${INPUTS_INITIAL_TOTAL_ROW}/B12,"—")'
    )

    admin_meta_row = pay_row + 1
    ws_m[f"A{admin_meta_row}"] = f"Roster hours — investor scale (ref.){H}"
    ws_m[f"B{admin_meta_row}"] = (
        f'=IF(OR(B{principal_a_div_row}<=0,INPUTS!$B${INPUTS_INVESTOR_NET_ROW}<=0),"—",'
        f"ROUNDUP(SUM(ROSTER!$C$2:$C$21)*(INPUTS!$B${INPUTS_INVESTOR_NET_ROW}/B{principal_a_div_row}),1))"
    )

    net_row = 12
    for addr in [
        "B3", "B4", "B5", "B6", "B7", "B8", "B9", "B10", "B11", "B12",
        "B16", "B17", "B18", "B19", "B20", "B21", "B22", "B23",
    ]:
        ws_m[addr].number_format = "#,##0.00"
    for addr in ["B13", "B14"]:
        ws_m[addr].number_format = "General"
    ws_m[f"B{pay_row}"].number_format = "General"
    ws_m[f"B{admin_meta_row}"].number_format = "General"

    ws_m.column_dimensions["A"].width = 52
    ws_m.column_dimensions["B"].width = 22

    ws_m.conditional_formatting.add(
        "B12",
        FormulaRule(formula=["B12<0"], stopIfTrue=True, fill=RED_CF),
    )
    ws_m.conditional_formatting.add(
        "B13",
        FormulaRule(
            formula=["$B$12<INPUTS!$B$6"],
            stopIfTrue=True,
            fill=RED_CF,
        ),
    )
    ws_m.conditional_formatting.add(
        "B8",
        FormulaRule(formula=["B8<0"], stopIfTrue=True, fill=RED_CF),
    )

    # --- SCENARIOS ---
    ws_s = wb.create_sheet("SCENARIOS", 8)
    ws_s["A1"] = f"Metric{H}"
    for col, val in enumerate([1, 2, 3, 5, 8], start=2):
        ws_s[f"{get_column_letter(col)}1"] = val
    ws_s["A1"].font = HEADER_FONT

    for col in range(2, 7):
        c = get_column_letter(col)
        ws_s[f"{c}2"] = (
            f"=IF({ROSTER_COUNT}=0,0,MODEL!$B$5*({c}$1/MAX({ROSTER_COUNT},1)))"
        )

    for col in range(2, 7):
        c = get_column_letter(col)
        ws_s[f"{c}3"] = (
            f"=IF({ROSTER_COUNT}=0,0,MODEL!$B$6*({c}$1/MAX({ROSTER_COUNT},1)))"
        )
        ws_s[f"{c}4"] = f"={c}2-{c}3-INPUTS!$B${INPUTS_PARTNER_DRAW_ROW}"
        ws_s[f"{c}5"] = f"=IF({c}4>0,{c}4*(1-INPUTS!$B${INPUTS_TAX_ROW}),{c}4)"
        ws_s[f"{c}6"] = f"={c}5*INPUTS!$B${INPUTS_DIVIDEND_ROW}"

    for col in range(2, 7):
        for row in range(2, 7):
            ws_s.cell(row=row, column=col).number_format = "#,##0.00"

    for r, lab in enumerate(
        [
            f"Clinic margin (scaled){H}",
            f"Operating costs (scaled){H}",
            f"Pre-tax after draw{H}",
            f"Net (tax approx){H}",
            f"Dividend pool (approx){H}",
        ],
        start=2,
    ):
        ws_s[f"A{r}"] = lab

    ws_s["A8"] = f"Clinic % sensitivity — net profit{H}"
    ws_s["A8"].font = Font(bold=True)
    ws_s["A9"] = f"Clinic %{H}"
    ws_s["B9"], ws_s["C9"], ws_s["D9"] = 0.33, 0.4, 0.5
    for col in range(2, 5):
        ws_s.cell(row=9, column=col).number_format = "0.00%"
    ws_s["A10"] = f"Net profit (approx){H}"
    for col_letter, _col in (("B", 2), ("C", 3), ("D", 4)):
        pbt = (
            f"{ROSTER_F_SUM}*{col_letter}9+"
            f"INPUTS!$B${INPUTS_DIR_CLIN_SESS_ROW}*INPUTS!$B${INPUTS_DIR_BILL_RATE_ROW}*{col_letter}9"
            f"-MODEL!$B$6-INPUTS!$B${INPUTS_PARTNER_DRAW_ROW}"
        )
        ws_s.cell(row=10, column=_col).value = (
            f"=IF(({pbt})>0,({pbt})*(1-INPUTS!$B${INPUTS_TAX_ROW}),({pbt}))"
        )
        ws_s.cell(row=10, column=_col).number_format = "#,##0.00"

    ws_s.column_dimensions["A"].width = 40
    for col in range(2, 7):
        ws_s.column_dimensions[get_column_letter(col)].width = 12

    ws_s.conditional_formatting.add(
        "B4:F6",
        FormulaRule(formula=["B4<0"], stopIfTrue=True, fill=RED_CF),
    )

    # --- SCENARIO_B_LIGHT ---
    ws_bl = wb.create_sheet("SCENARIO_B_LIGHT", 9)
    ws_bl["A1"] = f"Scenario B — light{H}"
    ws_bl["A1"].font = Font(bold=True, size=12)
    ws_bl["A2"] = f"Session price — low{H}"
    ws_bl["B2"] = 120
    ws_bl["A3"] = f"Session price — high{H}"
    ws_bl["B3"] = 150
    ws_bl["A4"] = f"Midpoint{H}"
    ws_bl["B4"] = "=(B2+B3)/2"
    ws_bl["A5"] = f"Clinic %{H}"
    ws_bl["B5"] = f"={clinic_ref}"
    ws_bl["A6"] = f"Revenue / session{H}"
    ws_bl["B6"] = "=B4*B5"
    ws_bl["A7"] = f"Monthly session-equiv (roster hours × 4){H}"
    ws_bl["B7"] = "=SUM(ROSTER!$C$2:$C$21)*4"
    ws_bl["A8"] = f"Monthly revenue{H}"
    ws_bl["B8"] = "=B7*B6"
    ws_bl["A9"] = f"Opex Path B{H}"
    ws_bl["B9"] = f"={opex_path_b}"
    ws_bl["A10"] = f"Supervision (excl.){H}"
    ws_bl["B10"] = 0
    ws_bl["A11"] = f"Admin flat{H}"
    ws_bl["B11"] = (
        f"=IF({ROSTER_COUNT}>=INPUTS!$B${INPUTS_PS_MIN_ROW},INPUTS!$B${INPUTS_ADMIN_FLAT_ROW},0)"
    )
    ws_bl["A12"] = f"Total costs{H}"
    ws_bl["B12"] = "=B9+B10+B11"
    ws_bl["A14"] = f"PBT{H}"
    ws_bl["B14"] = "=B8-B12"
    ws_bl["A15"] = f"Tax{H}"
    ws_bl["B15"] = f"=IF(B14>0,B14*INPUTS!$B${INPUTS_TAX_ROW},0)"
    ws_bl["A16"] = f"Net{H}"
    ws_bl["B16"] = "=B14-B15"
    ws_bl["A18"] = f"ROI vs initial{H}"
    ws_bl["B18"] = (
        f'=IF(INPUTS!$B${INPUTS_INITIAL_TOTAL_ROW}<=0,"—",(B16*12)/INPUTS!$B${INPUTS_INITIAL_TOTAL_ROW})'
    )
    for r, fmt in [(2, fmt_money), (3, fmt_money), (4, fmt_money), (6, fmt_money), (8, fmt_money), (16, fmt_money)]:
        ws_bl[f"B{r}"].number_format = fmt
    ws_bl["B18"].number_format = "0.00%"
    for r in (7, 9, 10, 11, 12, 14, 15):
        ws_bl[f"B{r}"].number_format = "#,##0.00"
    for addr in ("B2", "B3"):
        ws_bl[addr].fill = BLUE_INPUT

    # --- PARTNER_L_COVERAGE ---
    ws_lor = wb.create_sheet("PARTNER_L_COVERAGE", 10)
    ws_lor["A1"] = f"Partner_L — fixed draw vs modeled economics{H}"
    ws_lor["A1"].font = Font(bold=True, size=12)
    ws_lor["A3"] = f"Partner_L_Draw (target){H}"
    ws_lor["B3"] = f"=INPUTS!$B${INPUTS_PARTNER_DRAW_ROW}"
    ws_lor["A4"] = f"Net profit (MODEL){H}"
    ws_lor["B4"] = f"=MODEL!$B${net_row}"
    ws_lor["A5"] = f"Partner_L dividend (PARTNERSHIP_STRUCTURE){H}"
    ws_lor["B5"] = "=MODEL!$B$20"
    ws_lor["A6"] = f"Gap (draw − dividend){H}"
    ws_lor["B6"] = "=MAX(0,B3-B5)"
    for r in (3, 4, 5, 6):
        ws_lor[f"B{r}"].number_format = fmt_money

    # --- EXPENSES ---
    ws_e = wb.create_sheet("EXPENSES", 11)
    ws_e["A1"] = f"Category{H}"
    ws_e["B1"] = f"Monthly{H}"
    ws_e["C1"] = f"Annual{H}"
    ws_e["D1"] = f"Include?{H}"
    for c in ("A1", "B1", "C1", "D1"):
        ws_e[c].font = HEADER_FONT
    er = 2
    ws_e[f"A{er}"] = f"Director hybrid pay{H}"
    ws_e[f"B{er}"] = (
        f"=INPUTS!$B${INPUTS_DIR_ADMIN_HRS_ROW}*INPUTS!$B${INPUTS_DIR_ADMIN_RATE_ROW}+"
        f"INPUTS!$B${INPUTS_DIR_CLIN_SESS_ROW}*INPUTS!$B${INPUTS_DIR_CLIN_PAY_ROW}"
    )
    ws_e[f"D{er}"] = True
    er += 1
    ws_e[f"A{er}"] = f"Other variable (legacy + admin + supervision){H}"
    ws_e[f"B{er}"] = f"=MODEL!$B$6-OPERATING_COSTS!$B${oc_total_row}-B2"
    ws_e[f"D{er}"] = True
    er += 1
    for oc_r in range(oc_first, oc_last + 1):
        ws_e[f"A{er}"] = f"=OPERATING_COSTS!$A${oc_r}"
        ws_e[f"B{er}"] = f"=OPERATING_COSTS!$B${oc_r}"
        ws_e[f"D{er}"] = True
        er += 1
    ws_e[f"A{er}"] = f"Law 25 privacy [PENDING_DATA]{H}"
    ws_e[f"B{er}"] = 0
    ws_e[f"D{er}"] = True
    ws_e[f"B{er}"].fill = BLUE_INPUT
    law25_row = er
    er += 1
    last_exp = er - 1
    tot_exp_row = er
    for i in range(2, last_exp + 1):
        ws_e[f"C{i}"] = f"=B{i}*12"
    ws_e[f"A{tot_exp_row}"] = f"Total monthly{H}"
    ws_e[f"B{tot_exp_row}"] = f"=SUMIF(D2:D{last_exp},TRUE,B2:B{last_exp})"
    ws_e[f"A{tot_exp_row}"].font = Font(bold=True)
    ws_e[f"B{tot_exp_row}"].font = Font(bold=True)
    for i in range(2, tot_exp_row + 1):
        ws_e[f"B{i}"].number_format = "#,##0.00"
        if i <= last_exp:
            ws_e[f"C{i}"].number_format = "#,##0.00"

    # --- DASHBOARD ---
    ws_d = wb.create_sheet("DASHBOARD")
    ws_d.sheet_view.showGridLines = False
    ws_d["A2"] = f"Total clinic margin{H}"
    ws_d["B2"] = "=MODEL!B5"
    ws_d["A3"] = f"Net profit (final){H}"
    ws_d["B3"] = f"=MODEL!B{net_row}"
    ws_d["A4"] = f"Break-even gross (roster ref.){H}"
    ws_d["B4"] = "=MODEL!B17"
    ws_d["A5"] = f"Payback{H}"
    ws_d["B5"] = f"=MODEL!B{pay_row}"
    ws_d["A6"] = f"Dividend pool{H}"
    ws_d["B6"] = "=MODEL!B18"
    ws_d["A7"] = f"Investor scale (hours){H}"
    ws_d["B7"] = f"=MODEL!B{admin_meta_row}"
    ws_d["A8"] = f"Path B net{H}"
    ws_d["B8"] = "=SCENARIO_B_LIGHT!B16"
    ws_d["A9"] = f"Partner_L coverage gap{H}"
    ws_d["B9"] = "=PARTNER_L_COVERAGE!B6"
    ws_d["A10"] = f"Solvency{H}"
    ws_d["B10"] = "=MODEL!B14"
    ws_d["A11"] = f"Initial capital — valuation ({FILE_INITIAL_CAPITAL_VALUATION}){H}"
    ws_d["B11"] = f"=INPUTS!$B${INPUTS_IMPL_CAP_ROW}"
    ws_d["A12"] = f"ROI — initial capital (valuation){H}"
    ws_d["B12"] = (
        f'=IF(INPUTS!$B${INPUTS_IMPL_CAP_ROW}<=0,"—",'
        f"(MODEL!$B${net_row}*12)/INPUTS!$B${INPUTS_IMPL_CAP_ROW})"
    )
    ws_d["A13"] = f"Redemption alert{H}"
    ws_d["B13"] = "=MODEL!B13"

    corp = Font(name="Calibri Light", size=11, color="374151")
    corp_title = Font(name="Calibri Light", size=15, bold=True, color="111827")
    corp_head = Font(name="Calibri Light", size=10, bold=True, color="6B7280")
    fill_print = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")
    ws_d.merge_cells("A24:F24")
    ws_d["A24"] = f"Clinical Operations System — Master Dossier (Print View){H}"
    ws_d["A24"].font = corp_title
    ws_d["A24"].alignment = Alignment(horizontal="center", vertical="center")
    ws_d["A24"].fill = fill_print
    ws_d.row_dimensions[24].height = 28
    for r, (la, fmla) in enumerate(
        [
            ("Total clinic margin (CAD/mo)", "=MODEL!B5"),
            ("Net profit — final (CAD/mo)", "=MODEL!B12"),
            ("Partner_L_Draw (CAD/mo)", "=INPUTS!B6"),
            ("Redemption flag (MODEL)", "=MODEL!B13"),
            ("Initial capital — line-item audit (INPUTS)", "=INPUTS!B25"),
        ],
        start=25,
    ):
        ws_d[f"A{r}"] = la
        ws_d[f"B{r}"] = fmla
        ws_d[f"A{r}"].font = corp_head
        ws_d[f"B{r}"].font = corp
        ws_d[f"B{r}"].number_format = "#,##0.00" if r != 28 else "General"
        ws_d[f"A{r}"].fill = fill_print
        ws_d[f"B{r}"].fill = fill_print
    ws_d.print_options.horizontalCentered = True
    ws_d.page_setup.orientation = "portrait"
    ws_d.page_setup.fitToPage = True
    ws_d.page_setup.fitToWidth = 1
    ws_d.page_setup.fitToHeight = 100
    ws_d.print_area = "A24:F31"
    ws_d.oddFooter.left.text = (
        "Confidential — Proprietary — Clinical Operations System (COS) — The Architect"
    )
    ws_d.oddFooter.left.size = 9

    for r in range(2, 14):
        ws_d[f"A{r}"].font = Font(size=11, bold=True)
        ws_d[f"B{r}"].alignment = Alignment(horizontal="right")
    for r in (2, 3, 4, 6, 8, 9, 11):
        ws_d[f"B{r}"].number_format = "#,##0.00"
    ws_d["B5"].number_format = "General"
    ws_d["B7"].number_format = "General"
    ws_d["B10"].number_format = "General"
    ws_d["B12"].number_format = "0.00%"
    ws_d["B13"].number_format = "General"

    fill_disc = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    ws_d.merge_cells("M37:P37")
    ws_d["M37"] = "Input_Variables"
    ws_d["M37"].font = Font(name="Calibri Light", size=8, italic=True, color="9CA3AF")
    ws_d["M37"].alignment = Alignment(horizontal="right", vertical="center")
    ws_d["M38"] = "ADJUSTMENT_RESERVE"
    ws_d["M38"].font = Font(name="Calibri Light", size=8, color="6B7280")
    ws_d["N38"] = 0
    ws_d["N38"].fill = fill_disc
    ws_d["N38"].font = Font(name="Calibri Light", size=9, color="4B5563")
    ws_d["N38"].number_format = fmt_money
    ws_d.column_dimensions["M"].width = 24
    ws_d.column_dimensions["N"].width = 14

    pie_end = last_exp
    chart_line = LineChart()
    chart_line.title = f"Net (scenario){H}"
    chart_line.add_data(Reference(ws_s, min_col=2, min_row=5, max_col=6, max_row=5), titles_from_data=False)
    chart_line.set_categories(Reference(ws_s, min_col=2, min_row=1, max_col=6, max_row=1))
    ws_d.add_chart(chart_line, "H2")
    chart_bar = BarChart()
    chart_bar.type = "col"
    chart_bar.add_data(Reference(ws_s, min_col=2, min_row=2, max_col=6, max_row=2), titles_from_data=False)
    chart_bar.set_categories(Reference(ws_s, min_col=2, min_row=1, max_col=6, max_row=1))
    ws_d.add_chart(chart_bar, "H18")
    chart_pie = PieChart()
    chart_pie.add_data(Reference(ws_e, min_col=2, min_row=2, max_row=pie_end), titles_from_data=False)
    chart_pie.set_categories(Reference(ws_e, min_col=1, min_row=2, max_row=pie_end))
    chart_pie.dataLabels = DataLabelList()
    chart_pie.dataLabels.showPercent = True
    ws_d.add_chart(chart_pie, "H34")

    ws_in[f"B{INPUTS_ADJUSTMENT_RESERVE_ROW}"] = "=DASHBOARD!$N$38"
    ws_in[f"B{INPUTS_ADJUSTMENT_RESERVE_ROW}"].number_format = fmt_money
    ws_in[f"B{INPUTS_ADJUSTMENT_RESERVE_ROW}"].comment = _draft_comment(
        "Single editable control on DASHBOARD (N38); flows to MODEL via INPUTS."
    )

    # --- PRINT_SUMMARY (publisher) ---
    thin = Side(style="thin", color="CCCCCC")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    ps_title = Font(name="Calibri Light", size=18, bold=True, color="111827")
    ps_sec = Font(name="Calibri Light", size=12, bold=True, color="374151")
    ps_body = Font(name="Calibri Light", size=11, color="374151")
    ps_meta = Font(name="Calibri Light", size=9, color="6B7280")
    ws_ps = wb.create_sheet("PRINT_SUMMARY")
    ws_ps.sheet_view.showGridLines = False
    ws_ps.merge_cells("A1:F3")
    ws_ps["A1"] = "[PROJECT_ID: COS-2026-QC]"
    ws_ps["A1"].font = ps_title
    ws_ps["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for rr in range(1, 4):
        ws_ps.row_dimensions[rr].height = 18 if rr > 1 else 32
    for r in range(1, 4):
        for c in range(1, 7):
            ws_ps.cell(row=r, column=c).border = box
    ws_ps.merge_cells("A4:F4")
    ws_ps["A4"] = "Clinical Operations System (COS) — financial reporting extract"
    ws_ps["A4"].font = ps_meta
    ws_ps["A4"].alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 7):
        ws_ps.cell(row=4, column=c).border = box
    ws_ps.merge_cells("A5:F5")
    ws_ps["A5"] = "Executive summary — CAPEX, OpEx, solvency (linked)"
    ws_ps["A5"].font = ps_sec
    for c in range(1, 7):
        ws_ps.cell(row=5, column=c).border = box
    rows_ps = [
        ("CAPEX — Total Initial Valuation (CAPEX_SUMMARY)", "=CAPEX_SUMMARY!B3"),
        ("Monthly OpEx — operating detail (OPERATING_COSTS)", f"={opex_total}"),
        ("Total operating load incl. hybrid pay (MODEL)", "=MODEL!B6"),
        ("Net profit — final (MODEL)", "=MODEL!B12"),
        ("Solvency validation (MODEL)", "=MODEL!B14"),
        ("Redemption alert text (MODEL)", "=MODEL!B13"),
    ]
    for i, (la, fm) in enumerate(rows_ps, start=6):
        ws_ps[f"A{i}"] = la
        ws_ps[f"B{i}"] = fm
        ws_ps[f"A{i}"].font = ps_body
        ws_ps[f"B{i}"].font = ps_body
        ws_ps[f"B{i}"].number_format = "General" if i >= 10 else "#,##0.00"
        for c in range(1, 7):
            ws_ps.cell(row=i, column=c).border = box
    ws_ps.column_dimensions["A"].width = 48
    ws_ps.column_dimensions["B"].width = 28
    for col in range(3, 7):
        ws_ps.column_dimensions[get_column_letter(col)].width = 6
    ws_ps.print_options.horizontalCentered = True
    ws_ps.page_setup.orientation = "portrait"
    ws_ps.page_setup.fitToPage = True
    ws_ps.page_setup.fitToWidth = 1
    ws_ps.page_setup.fitToHeight = 100
    ws_ps.print_area = "A1:F11"
    ws_ps.oddFooter.left.text = (
        "Confidential - Proprietary - Clinical Operations System (COS) - The Architect"
    )
    ws_ps.oddFooter.left.size = 9

    wb.move_sheet(ws_d, offset=-1)

    dv_clinic = DataValidation(type="decimal", operator="between", formula1=0, formula2=1)
    dv_clinic.add(f"B{INPUTS_CLINIC_PCT_ROW}")
    dv_util = DataValidation(type="decimal", operator="between", formula1=0.5, formula2=1)
    dv_util.add(f"B{INPUTS_UTIL_ROW}")
    dv_div = DataValidation(type="decimal", operator="between", formula1=0, formula2=1)
    dv_div.add(f"B{INPUTS_DIVIDEND_ROW}")
    dv_dir_pct = DataValidation(type="decimal", operator="between", formula1=0, formula2=1)
    dv_dir_pct.add(f"B{INPUTS_DIR_PCT_ROW}")
    for dv in (dv_clinic, dv_util, dv_div, dv_dir_pct):
        ws_in.add_data_validation(dv)

    names = [
        ("OperatingCostsMonthly", opex_total),
        ("ClinicMargin", "MODEL!$B$5"),
        ("NetProfit", f"MODEL!$B${net_row}"),
        ("InitialCapitalContribution", f"INPUTS!$B${INPUTS_IMPL_CAP_ROW}"),
    ]
    for name, ref in names:
        wb.defined_names.add(DefinedName(name, attr_text=ref))
    wb.defined_names.add(DefinedName("Input_Variables", attr_text="DASHBOARD!$N$38"))

    input_unlock = []
    for r in range(2, INPUTS_INITIAL_TOTAL_ROW + 1):
        if r in (
            INPUTS_INITIAL_TOTAL_ROW,
            INPUTS_WEEKS_FACTOR_ROW,
            INPUTS_DIR_CLIN_PAY_ROW,
            INPUTS_ADJUSTMENT_RESERVE_ROW,
        ):
            continue
        input_unlock.append(f"B{r}")
    cap_unlock = ["B2"]
    for r in range(cap_first, cap_last + 1):
        cap_unlock.extend([f"A{r}", f"B{r}", f"C{r}", f"D{r}"])
    oc_unlock = [f"B{r}" for r in range(oc_first, oc_last + 1)]
    av_unlock: list[str] = []
    for r in range(av_first, av_last_data + 1):
        av_unlock.extend([f"A{r}", f"B{r}", f"C{r}"])
    roster_unlock = [f"{c}{r}" for r in range(roster_first, roster_last + 1) for c in "ABCD"]
    tier_unlock = [f"A{r}" for r in range(2, 5)] + [f"B{r}" for r in range(2, 5)]
    bl_unlock = ["B2", "B3"]
    exp_unlock = [f"B{law25_row}"]
    dashboard_unlock = [DASHBOARD_ADJ_CELL]

    for ws in (ws_in, ws_ro, ws_tier, ws_oc, ws_av, ws_capex, ws_cap, ws_m, ws_s, ws_bl, ws_lor, ws_e, ws_d, ws_ps):
        _lock_all_cells(ws)
    _unlock_cells(ws_in, input_unlock)
    _unlock_cells(ws_ro, roster_unlock)
    _unlock_cells(ws_tier, tier_unlock)
    _unlock_cells(ws_oc, oc_unlock)
    _unlock_cells(ws_cap, cap_unlock)
    _unlock_cells(ws_av, av_unlock)
    _unlock_cells(ws_bl, bl_unlock)
    _unlock_cells(ws_e, exp_unlock)
    _unlock_cells(ws_d, dashboard_unlock)

    _pw = hash_password(WORKBOOK_SHEET_PASSWORD)
    for ws in (ws_in, ws_ro, ws_tier, ws_oc, ws_av, ws_capex, ws_cap, ws_m, ws_s, ws_bl, ws_lor, ws_e, ws_d, ws_ps):
        ws.protection.sheet = True
        ws.protection.password = _pw
        ws.protection.formatCells = False
        ws.protection.formatColumns = False
        ws.protection.formatRows = False
        ws.protection.insertColumns = False
        ws.protection.insertRows = False
        ws.protection.deleteColumns = False
        ws.protection.deleteRows = False
        ws.protection.selectLockedCells = True
        ws.protection.selectUnlockedCells = True

    _export_model_sheet_csv(wb, base / "output" / "MODEL.csv")
    wb.save(out)
    return out


def main() -> None:
    from .health_check import run_health_check

    out = build_workbook()
    print(out)
    run_health_check()


if __name__ == "__main__":
    main()
