"""
Operational health checks for ANALISIS_FINANCIERO_PRE_LAUNCH.xlsx.
Run after build or on startup: python -m analista_financiero_clinica_virtual.health_check
"""

from __future__ import annotations

import csv
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


def project_root() -> Path:
    return Path(__file__).resolve().parent.parent


def default_workbook_path() -> Path:
    return project_root() / "output" / "ANALISIS_FINANCIERO_PRE_LAUNCH.xlsx"


def load_numeric(wb, sheet: str, cell: str) -> float | None:
    ws = wb[sheet]
    v = ws[cell].value
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        if v.strip() in ("—", "-", ""):
            return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _inp(wb, row: int) -> float:
    """INPUTS column B — must match build_prelaunch_workbook row indices."""
    v = load_numeric(wb, "INPUTS", f"B{row}")
    return float(v) if v is not None else 0.0


def synthetic_net_profit_v3(wb) -> float | None:
    """Reserved: full MODEL replication without Excel calc cache (optional extension)."""
    return None


def run_health_check(xlsx_path: Path | None = None) -> int:
    """Return 0 if no blocking issues; prints alerts to stdout."""
    path = xlsx_path or default_workbook_path()
    if not path.is_file():
        print(f"ALERT: Workbook not found: {path}", file=sys.stderr)
        return 2

    wb = load_workbook(path, data_only=True)

    np_ = load_numeric(wb, "MODEL", "B12")
    if np_ is None:
        np_ = synthetic_net_profit_v3(wb)

    # Partner_L dividend share (MODEL B20) vs final net (B12)
    lor_div = load_numeric(wb, "MODEL", "B20")
    if lor_div is None:
        lor_div = 0.0

    roster_gross = load_numeric(wb, "ROSTER", "F22")
    be_gross_need = load_numeric(wb, "MODEL", "B17")

    exit_code = 0
    if np_ is not None and np_ < 0 and lor_div > 0:
        print(
            "INSOLVENCY RISK: Partner draw exceeds profit. "
            "(Net profit < 0 and Partner_L dividend > 0.)"
        )
        exit_code = 1

    if (
        roster_gross is not None
        and be_gross_need is not None
        and be_gross_need > 0
        and roster_gross < be_gross_need
    ):
        print(
            "VOLUME ALERT: Roster gross billing below approximate break-even gross. "
            "(Increase weekly hours or rates.)"
        )
        if exit_code == 0:
            exit_code = 1

    if exit_code == 0:
        print("Health check: OK (no alerts).")

    return exit_code


def print_missing_data_report() -> None:
    """Static audit of generator source: PENDING_INPUT and placeholder numerics."""
    root = project_root()
    src = root / "analista_financiero_clinica_virtual" / "build_prelaunch_workbook.py"
    text = src.read_text(encoding="utf-8")
    pending_lines = [ln.strip() for ln in text.splitlines() if "ST_PENDING" in ln and "ST_PENDING =" not in ln]
    mock_nums = []
    for pat in (r"\b300\b", r"\b400\b", r"\b80000\b", r"\b36000\b"):
        if re.search(pat, text):
            mock_nums.append(pat)

    print("=== MISSING DATA REPORT (generator source) ===")
    print(f"Source: {src}")
    print(f"Rows with ST_PENDING assignment: {len(pending_lines)}")
    for ln in pending_lines[:40]:
        print(f"  - {ln[:120]}")
    if len(pending_lines) > 40:
        print(f"  ... ({len(pending_lines) - 40} more)")
    print("Hard-coded placeholder scan (300/400 municipal & SW annual; 80k/36k):")
    if re.search(r"\b300\b", text):
        print("  - 300 appears (Municipal annual placeholder in INPUTS).")
    if re.search(r"\b400\b", text):
        print("  - 400 appears (SW liability annual placeholder in INPUTS).")
    if re.search(r"80000|80,?000", text):
        print("  - 80k pattern found in source.")
    else:
        print("  - No 80k mock pattern in generator.")
    if re.search(r"36000|36,?000", text):
        print("  - 36k pattern found in source.")
    else:
        print("  - No 36k mock pattern in generator.")
    print("=== END REPORT ===\n")


def implementation_debt_from_csv() -> float:
    csv_path = project_root() / "INFRASTRUCTURE_ASSETS.csv"
    if not csv_path.is_file():
        csv_path = project_root() / "SHANE_ASSETS.csv"
    if not csv_path.is_file():
        return 0.0
    total = 0.0
    with csv_path.open(newline="", encoding="utf-8") as f:
        r = csv.DictReader(f)
        if not r.fieldnames:
            return 0.0
        key = "Total_Value_CAD" if "Total_Value_CAD" in r.fieldnames else "Total_Value"
        if key not in r.fieldnames:
            return 0.0
        for row in r:
            try:
                total += float(row.get(key) or 0)
            except (TypeError, ValueError):
                pass
    return total


def main() -> None:
    if "--report" in sys.argv:
        print_missing_data_report()
        return
    code = run_health_check()
    sys.exit(code)


if __name__ == "__main__":
    main()
